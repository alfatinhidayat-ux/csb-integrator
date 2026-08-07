import argparse
import os
import sys

import openpyxl
import pymysql

sys.path.insert(0, os.getcwd())

from config import Config
from db import DatabaseManager

CABANG_NAME_TO_ID = {
    "Kobisonta": 1,
    "Bula": 2,
    "Mandiri": 4,
    "Kairatu": 5,
    "Piru": 7,
}

HEADER_TABLE = "pos_ok"
DETAIL_TABLE = "pos_ok_detail"


def ensure_tables(db: DatabaseManager):
    print(f"Ensuring tables exist ({HEADER_TABLE}, {DETAIL_TABLE})...")
    with db.conn.cursor() as cur:
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS `{HEADER_TABLE}` (
                id BIGINT NOT NULL,
                cabang_id INT NOT NULL,
                no INT NULL,
                tanggal DATE NULL,
                no_nota VARCHAR(100) NULL,
                id_nota BIGINT NULL,
                cabang_nama VARCHAR(100) NULL,
                pelanggan VARCHAR(255) NULL,
                keterangan TEXT NULL,
                status_dokumen VARCHAR(50) NULL,
                total_biaya DECIMAL(18,2) NULL,
                bayar DECIMAL(18,2) NULL,
                tunai DECIMAL(18,2) NULL,
                qris_barcode DECIMAL(18,2) NULL,
                transfer DECIMAL(18,2) NULL,
                kartu_edc_kanal DECIMAL(18,2) NULL,
                kanal_lain DECIMAL(18,2) NULL,
                total_bayar DECIMAL(18,2) NULL,
                selisih DECIMAL(18,2) NULL,
                jml_baris_bayar INT NULL,
                kombinasi_kanal VARCHAR(255) NULL,
                bank_transfer VARCHAR(255) NULL,
                nama_transfer VARCHAR(255) NULL,
                kartu_jenis VARCHAR(100) NULL,
                kartu_edc VARCHAR(100) NULL,
                kartu_no VARCHAR(100) NULL,
                batal_keterangan TEXT NULL,
                batal_tanggal DATETIME NULL,
                batal_oleh VARCHAR(100) NULL,
                synced_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (id, cabang_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS `{DETAIL_TABLE}` (
                id BIGINT NOT NULL,
                cabang_id INT NOT NULL,
                pos_id BIGINT NOT NULL,
                no INT NULL,
                tanggal DATE NULL,
                no_nota VARCHAR(255) NULL,
                cabang_nama VARCHAR(100) NULL,
                status_nota VARCHAR(50) NULL,
                total_nota DECIMAL(18,2) NULL,
                produk_id INT NULL,
                kode_produk VARCHAR(100) NULL,
                nama_produk VARCHAR(255) NULL,
                sku VARCHAR(100) NULL,
                grup INT NULL,
                sub_grup INT NULL,
                merek VARCHAR(100) NULL,
                satuan VARCHAR(100) NULL,
                kode_satuan VARCHAR(50) NULL,
                jumlah DECIMAL(18,4) NULL,
                harga_satuan DECIMAL(18,2) NULL,
                diskon_persen DECIMAL(6,2) NULL,
                diskon_rp DECIMAL(18,2) NULL,
                subtotal DECIMAL(18,2) NULL,
                jml_retur DECIMAL(18,4) NULL,
                synced_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (id, cabang_id),
                KEY idx_pos_id (pos_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
    db.conn.commit()
    print("Tables ready.")


def iter_data_rows(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if any(v is not None for v in row):
            yield row


def build_header_records(ws):
    records = []
    for row in iter_data_rows(ws):
        try:
            id_val = int(row[3])
        except (TypeError, ValueError):
            id_val = 0
        records.append({
            "id": id_val,
            "cabang_id": resolve_bang_id_header(row),
            "no": row[0] if len(row) > 0 else None,
            "tanggal": row[1] if len(row) > 1 else None,
            "no_nota": row[2] if len(row) > 2 else None,
            "id_nota": row[3] if len(row) > 3 else None,
            "cabang_nama": row[5] if len(row) > 5 else None,
            "pelanggan": row[6] if len(row) > 6 else None,
            "keterangan": row[7] if len(row) > 7 else None,
            "status_dokumen": row[8] if len(row) > 8 else None,
            "total_biaya": row[9] if len(row) > 9 else None,
            "bayar": row[10] if len(row) > 10 else None,
            "tunai": row[11] if len(row) > 11 else None,
            "qris_barcode": row[12] if len(row) > 12 else None,
            "transfer": row[13] if len(row) > 13 else None,
            "kartu_edc_kanal": row[14] if len(row) > 14 else None,
            "kanal_lain": row[15] if len(row) > 15 else None,
            "total_bayar": row[16] if len(row) > 16 else None,
            "selisih": row[17] if len(row) > 17 else None,
            "jml_baris_bayar": row[18] if len(row) > 18 else None,
            "kombinasi_kanal": row[19] if len(row) > 19 else None,
            "bank_transfer": row[20] if len(row) > 20 else None,
            "nama_transfer": row[21] if len(row) > 21 else None,
            "kartu_jenis": row[22] if len(row) > 22 else None,
            "kartu_edc": row[23] if len(row) > 23 else None,
            "kartu_no": row[24] if len(row) > 24 else None,
            "batal_keterangan": row[25] if len(row) > 25 else None,
            "batal_tanggal": row[26] if len(row) > 26 else None,
            "batal_oleh": row[27] if len(row) > 27 else None,
        })
    return records


def build_detail_records(ws):
    records = []
    for row in iter_data_rows(ws):
        try:
            id_val = int(row[7])
        except (TypeError, ValueError):
            id_val = 0
        try:
            pos_id = int(row[3])
        except (TypeError, ValueError):
            pos_id = 0
        records.append({
            "id": id_val,
            "cabang_id": resolve_bang_id_detail(row),
            "pos_id": pos_id,
            "no": row[0] if len(row) > 0 else None,
            "tanggal": row[1] if len(row) > 1 else None,
            "no_nota": row[2] if len(row) > 2 else None,
            "cabang_nama": row[4] if len(row) > 4 else None,
            "status_nota": row[5] if len(row) > 5 else None,
            "total_nota": row[6] if len(row) > 6 else None,
            "produk_id": row[8] if len(row) > 8 else None,
            "kode_produk": row[9] if len(row) > 9 else None,
            "nama_produk": row[10] if len(row) > 10 else None,
            "sku": row[11] if len(row) > 11 else None,
            "grup": row[12] if len(row) > 12 else None,
            "sub_grup": row[13] if len(row) > 13 else None,
            "merek": row[14] if len(row) > 14 else None,
            "satuan": row[15] if len(row) > 15 else None,
            "kode_satuan": row[16] if len(row) > 16 else None,
            "jumlah": row[17] if len(row) > 17 else None,
            "harga_satuan": row[18] if len(row) > 18 else None,
            "diskon_persen": row[19] if len(row) > 19 else None,
            "diskon_rp": row[20] if len(row) > 20 else None,
            "subtotal": row[21] if len(row) > 21 else None,
            "jml_retur": row[22] if len(row) > 22 else None,
        })
    return records


def resolve_bang_id_header(row):
    raw = row[4] if len(row) > 4 else None
    if raw is not None:
        try:
            return int(raw)
        except (TypeError, ValueError):
            pass
    name = row[5] if len(row) > 5 else None
    if name and name in CABANG_NAME_TO_ID:
        return CABANG_NAME_TO_ID[name]
    return 1


def resolve_bang_id_detail(row):
    name = row[4] if len(row) > 4 else None
    if name and name in CABANG_NAME_TO_ID:
        return CABANG_NAME_TO_ID[name]
    return 1


def upsert_records(db, table, records, chunk_size=1000):
    if not records:
        return
    cols = list(records[0].keys())
    col_names = ", ".join(f"`{c}`" for c in cols)
    placeholders = ", ".join(["%s"] * len(cols))
    update_cols = [c for c in cols if c not in ("id", "cabang_id")]
    update_clause = ", ".join(f"`{c}` = VALUES(`{c}`)" for c in update_cols)
    sql = f"INSERT INTO `{table}` ({col_names}) VALUES ({placeholders}) ON DUPLICATE KEY UPDATE {update_clause}"

    total = len(records)
    for start in range(0, total, chunk_size):
        chunk = records[start:start + chunk_size]
        batch = [[rec.get(c) for c in cols] for rec in chunk]
        for attempt in range(3):
            try:
                db.reconnect()
                with db.conn.cursor() as cur:
                    cur.executemany(sql, batch)
                db.conn.commit()
                break
            except (pymysql.err.OperationalError, pymysql.err.InterfaceError):
                try:
                    db.conn.rollback()
                except Exception:
                    pass
                if attempt == 2:
                    raise
                print(f"    DB connection lost on {table} chunk {start}, reconnecting ({attempt + 1}/3)...")
                db.close()
                db.connect()
        print(f"    Upsert {table}: {min(start + chunk_size, total)}/{total} rows")


def main():
    parser = argparse.ArgumentParser(
        description="Sync POS Header & Detail from an Excel file into csb_db"
    )
    parser.add_argument(
        "-f", "--file",
        default="Penjualan_POS_Brighter_per_Cabang.xlsx",
        help="Path to the Excel source file",
    )
    parser.add_argument(
        "-e", "--env",
        action="store_true",
        help="Load configuration from environment variables",
    )
    parser.add_argument(
        "--only-header",
        action="store_true",
        help="Only sync the header sheets (skip '- Barang' sheets)",
    )
    args = parser.parse_args()

    config = Config.from_env()
    db = DatabaseManager(config, target_db="csb")
    db.connect()
    ensure_tables(db)

    print(f"Loading workbook: {args.file}")
    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)

    header_sheets = []
    detail_sheets = []
    for name in wb.sheetnames:
        if name.endswith(" - Barang"):
            detail_sheets.append(name)
        elif name not in ("Info", "Ringkasan"):
            header_sheets.append(name)

    print(f"Header sheets: {header_sheets}")
    print(f"Detail sheets: {detail_sheets}")

    total_headers = 0
    total_details = 0

    for name in header_sheets:
        ws = wb[name]
        print(f"\n--- Header sheet: {name} ---")
        records = build_header_records(ws)
        print(f" -> {len(records)} header rows")
        upsert_records(db, HEADER_TABLE, records)
        total_headers += len(records)

    if not args.only_header:
        for name in detail_sheets:
            ws = wb[name]
            print(f"\n--- Detail sheet: {name} ---")
            records = build_detail_records(ws)
            print(f" -> {len(records)} detail rows")
            upsert_records(db, DETAIL_TABLE, records)
            total_details += len(records)

    wb.close()
    db.close()

    print("\n" + "=" * 50)
    print(f"SYNC COMPLETE — {HEADER_TABLE}: {total_headers} rows, {DETAIL_TABLE}: {total_details} rows")
    print("=" * 50)


if __name__ == "__main__":
    main()